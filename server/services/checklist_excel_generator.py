import os
import pandas as pd
import logging
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

logger = logging.getLogger(__name__)


class ChecklistExcelGenerator:
    """
    Generate checklist Excel files from spec extraction results.
    Creates a structured checklist format similar to the reference file.
    """

    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def generate_checklist_excel(self, spec_result: Dict[str, Any], filename: str = None) -> str:
        """
        Generate a checklist Excel file from spec extraction results.

        Args:
            spec_result: The spec extraction result data
            filename: Optional filename for the Excel file

        Returns:
            Path to the generated Excel file
        """
        logger.info("=== GENERATING CHECKLIST EXCEL ===")

        # Create workbook and worksheet
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "チェックリスト"

        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"特記仕様書_チェックリスト_{timestamp}.xlsx"

        # Create the checklist structure
        self._create_header()
        self._create_basic_info_section(spec_result)
        self._create_estimate_info_section(spec_result)
        self._create_management_fee_section(spec_result)
        self._create_spec_sections(spec_result)

        # Save the file
        file_path = os.path.join("temp", filename)
        os.makedirs("temp", exist_ok=True)
        self.workbook.save(file_path)

        logger.info(f"Checklist Excel generated: {file_path}")
        return file_path

    def _create_header(self):
        """Create the header section of the checklist."""
        # Title
        self.worksheet['A1'] = "特記仕様書 チェックリスト"
        self.worksheet['A1'].font = Font(size=16, bold=True)
        self.worksheet.merge_cells('A1:H1')
        self.worksheet['A1'].alignment = Alignment(horizontal='center')

        # Generation date
        current_time = datetime.now()
        date_str = f"生成日時: {current_time.year}年{current_time.month}月{current_time.day}日 {current_time.hour:02d}:{current_time.minute:02d}:{current_time.second:02d}"
        self.worksheet['A2'] = date_str
        self.worksheet['A2'].font = Font(size=10)
        self.worksheet.merge_cells('A2:H2')

        # Add some spacing
        self.worksheet.row_dimensions[3].height = 20

    def _create_basic_info_section(self, spec_result: Dict[str, Any]):
        """Create the basic information section."""
        current_row = 4

        # Section header
        self.worksheet[f'A{current_row}'] = "基本情報"
        self.worksheet[f'A{current_row}'].font = Font(size=14, bold=True)
        self.worksheet[f'A{current_row}'].fill = PatternFill(
            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        self.worksheet.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        # File information
        self.worksheet[f'A{current_row}'] = "ファイル情報"
        self.worksheet[f'A{current_row}'].font = Font(bold=True)
        self.worksheet.merge_cells(f'A{current_row}:B{current_row}')

        self.worksheet[f'C{current_row}'] = "特記仕様書ファイル"
        self.worksheet[f'C{current_row}'].font = Font(bold=True)
        self.worksheet[f'D{current_row}'] = spec_result.get(
            'spec_filename', 'N/A')
        current_row += 1

        self.worksheet[f'C{current_row}'] = "見積積算参考資料ファイル"
        self.worksheet[f'C{current_row}'].font = Font(bold=True)
        self.worksheet[f'D{current_row}'] = spec_result.get(
            'estimate_filename', 'N/A')
        current_row += 2

    def _create_estimate_info_section(self, spec_result: Dict[str, Any]):
        """Create the estimate information section."""
        current_row = self.worksheet.max_row + 1

        # Section header
        self.worksheet[f'A{current_row}'] = "見積積算参考資料情報"
        self.worksheet[f'A{current_row}'].font = Font(size=14, bold=True)
        self.worksheet[f'A{current_row}'].fill = PatternFill(
            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        self.worksheet.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        # Estimate information fields
        estimate_fields = [
            ("省庁", spec_result.get('省庁', 'N/A')),
            ("年度", spec_result.get('年度', 'N/A')),
            ("経費工種", spec_result.get('経費工種', 'N/A')),
            ("施工地域工事場所", spec_result.get('施工地域工事場所', 'N/A'))
        ]

        for field_name, field_value in estimate_fields:
            self.worksheet[f'A{current_row}'] = field_name
            self.worksheet[f'A{current_row}'].font = Font(bold=True)
            self.worksheet[f'B{current_row}'] = field_value
            current_row += 1

        current_row += 1

    def _create_management_fee_section(self, spec_result: Dict[str, Any]):
        """Create the management fee section."""
        management_fee_subtables = spec_result.get(
            'management_fee_subtables', [])

        if not management_fee_subtables:
            return

        current_row = self.worksheet.max_row + 1

        # Section header
        self.worksheet[f'A{current_row}'] = "管理費区分：０以外"
        self.worksheet[f'A{current_row}'].font = Font(size=14, bold=True)
        self.worksheet[f'A{current_row}'].fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.worksheet.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        # Table headers
        headers = ["項目名", "参考番号", "摘要"]
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=current_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Table data
        for item in management_fee_subtables:
            self.worksheet[f'A{current_row}'] = item.get('item_name', '')
            self.worksheet[f'B{current_row}'] = item.get(
                'reference_number', '')
            self.worksheet[f'C{current_row}'] = item.get('notes', '')
            current_row += 1

        current_row += 1

    def _create_spec_sections(self, spec_result: Dict[str, Any]):
        """Create sections for each spec article."""
        sections = spec_result.get('sections', [])

        for section in sections:
            section_name = section.get('section', '')
            section_data = section.get('data', {})

            if not section_data:
                continue

            current_row = self.worksheet.max_row + 1

            # Section header
            self.worksheet[f'A{current_row}'] = section_name
            self.worksheet[f'A{current_row}'].font = Font(size=14, bold=True)
            self.worksheet[f'A{current_row}'].fill = PatternFill(
                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            self.worksheet.merge_cells(f'A{current_row}:H{current_row}')
            current_row += 1

            # Process section data
            self._add_section_data(section_data, current_row)

    def _add_section_data(self, data: Any, start_row: int):
        """Add section data to the worksheet."""
        current_row = start_row

        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, list) and len(value) > 0:
                    # Handle list data (tables)
                    self._add_table_data(key, value, current_row)
                    current_row = self.worksheet.max_row + 2
                else:
                    # Handle simple key-value pairs
                    self.worksheet[f'A{current_row}'] = key
                    self.worksheet[f'A{current_row}'].font = Font(bold=True)
                    self.worksheet[f'B{current_row}'] = str(
                        value) if value is not None else ''
                    current_row += 1
        elif isinstance(data, list):
            # Handle list data directly
            self._add_table_data("データ", data, current_row)
        else:
            # Handle simple values
            self.worksheet[f'A{current_row}'] = str(
                data) if data is not None else ''
            current_row += 1

    def _add_table_data(self, title: str, table_data: List[Dict], start_row: int):
        """Add table data to the worksheet."""
        if not table_data or not isinstance(table_data[0], dict):
            return

        current_row = start_row

        # Add title
        self.worksheet[f'A{current_row}'] = title
        self.worksheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

        # Get column headers from first row
        headers = list(table_data[0].keys())

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=current_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Add data rows
        for row_data in table_data:
            for col, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                self.worksheet.cell(row=current_row, column=col, value=value)
            current_row += 1

    def _apply_borders(self):
        """Apply borders to the worksheet."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border

    def _auto_adjust_columns(self):
        """Auto-adjust column widths."""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
