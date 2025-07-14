from typing import List, Dict, Any, Optional
from io import BytesIO
import logging
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border
import numpy as np
import tempfile

from ..schemas.tender import TenderItem, SubtableItem

# Import SubtableItem using absolute path to avoid import issues
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

# Debug: Check SubtableItem at import time
# print(f"DEBUG: SubtableItem type at import: {type(SubtableItem)}")
# print(
#     f"DEBUG: SubtableItem module at import: {getattr(SubtableItem, '__module__', None)}")
# print(
#     f"DEBUG: SubtableItem fields at import: {getattr(SubtableItem, '__fields__', None)}")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelTableExtractorService:
    """
    Service class that directly uses the excel_table_extractor_corrected.py
    for main table extraction in the backend.
    """

    def __init__(self):
        pass

    def extract_main_table_from_buffer(self, excel_buffer: BytesIO, sheet_name: str) -> List[TenderItem]:
        """
        Extract main table from Excel buffer using the standalone corrected extraction logic.
        This method directly uses the excel_table_extractor_corrected.py file.

        Args:
            excel_buffer: BytesIO buffer containing Excel data
            sheet_name: Name of the sheet to extract from

        Returns:
            List of TenderItem objects extracted from the main table
        """
        try:
            # Save buffer to temporary file for processing
            import tempfile

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(excel_buffer.getvalue())
                temp_file_path = temp_file.name

            try:
                # Use the standalone corrected extractor
                tender_items = self._extract_using_standalone_logic(
                    temp_file_path, sheet_name)

                logger.info(
                    f"Extracted {len(tender_items)} items from main table using standalone logic")
                return tender_items

            finally:
                # Clean up temporary file
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)

        except Exception as e:
            logger.error(f"Error extracting main table: {e}")
            # Do NOT fall back to original logic for main table extraction
            raise RuntimeError(f"Main table extraction failed: {e}")

    def extract_subtables(self, file_content: bytes, main_table_items: List[TenderItem]) -> List[SubtableItem]:
        """
        Extract subtables from Excel file by processing sheets sequentially.
        1. Extract reference numbers from main table's 摘要 column
        2. Find all subtable sheets and their reference numbers
        3. Process each sheet with its available references
        """
        all_subtable_items = []

        # Create temporary file for processing
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(file_content)
            tmp_file_path = tmp_file.name

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(tmp_file_path, data_only=True)

            # Step 1: Extract reference numbers from main table's 摘要 column
            main_references = self._extract_reference_numbers_from_main_table(
                main_table_items)
            logger.info(
                f"Found {len(main_references)} references from main table: {main_references[:10]}...")

            # Step 2: Find all subtable sheets and their reference numbers
            # Skip main sheet (index 0)
            sheets_to_process = workbook.sheetnames[1:]

            for sheet_index, sheet_name in enumerate(sheets_to_process):
                logger.info(
                    f"Processing sheet {sheet_index + 2}: {sheet_name}")

                # Get all reference numbers available in this sheet
                sheet_references = self._get_all_references_in_sheet(
                    tmp_file_path, sheet_name)

                if not sheet_references:
                    logger.info(
                        f"No reference numbers found in {sheet_name}, skipping")
                    continue

                logger.info(
                    f"Found {len(sheet_references)} references in {sheet_name}: {sheet_references[:10]}...")

                # Extract subtables from current sheet using its available references
                sheet_items = self._extract_subtables_from_sheet(
                    tmp_file_path, sheet_name, sheet_references)
                all_subtable_items.extend(sheet_items)

                logger.info(
                    f"Extracted {len(sheet_items)} items from {sheet_name}")

            workbook.close()

        except Exception as e:
            logger.error(f"Error extracting subtables: {e}")
        finally:
            import os
            os.unlink(tmp_file_path)

        logger.info(
            f"Total subtable items extracted: {len(all_subtable_items)}")
        return all_subtable_items

    def extract_subtables_from_buffer(self, excel_buffer: BytesIO, main_sheet_name: str, main_table_items: List[TenderItem]) -> List[SubtableItem]:
        """
        Wrapper method for API compatibility.
        Extract subtables from Excel buffer by processing sheets sequentially.
        """
        return self.extract_subtables(excel_buffer.getvalue(), main_table_items)

    def _extract_reference_numbers_from_main_table(self, main_table_items: List[TenderItem]) -> List[str]:
        """
        Extract reference numbers from main table items by looking at the 摘要 column.
        Extract patterns like 単1号, 代8号, 内1号, etc. (kanji + number + 号)
        """
        reference_numbers = set()

        # Extract reference numbers from main table items
        for item in main_table_items:
            if hasattr(item, 'raw_fields') and item.raw_fields:
                # Check specifically the 摘要 field
                tekiyo_value = item.raw_fields.get('摘要', '')
                if tekiyo_value and isinstance(tekiyo_value, str):
                    # Look for patterns like 単1号, 代8号, 内1号, etc.
                    import re
                    # Pattern: one or more kanji characters + digits + 号
                    matches = re.findall(r'[一-龯]+\d+号', tekiyo_value)
                    for match in matches:
                        reference_numbers.add(match)

                # Also check other fields that might contain reference numbers
                for field_name, field_value in item.raw_fields.items():
                    if field_value and isinstance(field_value, str):
                        # Look for reference patterns in all fields
                        import re
                        matches = re.findall(r'[一-龯]+\d+号', field_value)
                        for match in matches:
                            reference_numbers.add(match)

        sorted_references = sorted(list(reference_numbers))
        logger.info(
            f"Extracted reference numbers from main table: {sorted_references}")
        return sorted_references

    def _get_all_references_in_sheet(self, file_path: str, sheet_name: str) -> List[str]:
        """
        Extract all reference numbers found in a specific sheet.
        """
        from openpyxl import load_workbook

        reference_numbers = set()

        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook[sheet_name]

            # Search for reference patterns in the sheet
            # Increased search range
            max_rows = min(worksheet.max_row + 1, 3000)
            max_cols = min(worksheet.max_column + 1, 25)

            for row in range(1, max_rows):
                for col in range(1, max_cols):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        # Look for 単X号 patterns
                        import re
                        matches = re.findall(r'単\d+号', cell_value)
                        reference_numbers.update(matches)

                        # Also look for other patterns like 内X号, 代X号, etc.
                        other_matches = re.findall(
                            r'[内代工材][0-9]+号', cell_value)
                        reference_numbers.update(other_matches)

            workbook.close()

        except Exception as e:
            logger.error(
                f"Error extracting reference numbers from sheet {sheet_name}: {e}")

        result = sorted(list(reference_numbers), key=lambda x: int(
            re.search(r'\d+', x).group()) if re.search(r'\d+', x) else 0)
        return result

    def _find_header_row(self, extractor, ref_row: int) -> Optional[int]:
        """
        Find the header row for a subtable.
        It looks for a row that contains at least 2 keywords from ["名", "規", "注", "単", "数", "金", "摘"]
        and is not empty.
        """
        expected_keywords = ["名", "規", "注", "単", "数", "金", "摘"]

        # Look for header in the next few rows after the reference
        for row in range(ref_row + 1, min(ref_row + 10, extractor.worksheet.max_row + 1)):
            row_text = ""
            for col in range(1, extractor.worksheet.max_column + 1):
                cell_value = extractor.get_cell_value(row, col)
                if cell_value:
                    row_text += str(cell_value) + " "

            hit_count = sum(
                1 for keyword in expected_keywords if keyword in row_text)

            if hit_count >= 2 and row_text.strip():
                return row
        return None

    def _find_subtable_sheets(self, file_path: str, reference_numbers: List[str]) -> List[str]:
        """
        Find all sheets that contain reference numbers (subtable sheets).
        """
        from openpyxl import load_workbook

        subtable_sheets = []

        try:
            workbook = load_workbook(file_path, data_only=True)

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Check if this sheet contains any reference numbers
                # Search more thoroughly - some references might be deeper in the sheet
                found_refs = 0
                # Increased from 100 to 2000
                max_rows = min(worksheet.max_row + 1, 2000)
                # Increased from 10 to 25
                max_cols = min(worksheet.max_column + 1, 25)

                for row in range(1, max_rows):
                    for col in range(1, max_cols):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            for ref in reference_numbers:
                                if ref in cell_value:
                                    found_refs += 1
                                    break
                    if found_refs > 0:
                        break

                if found_refs > 0:
                    subtable_sheets.append(sheet_name)
                    logger.info(f"Found subtable sheet: {sheet_name}")

            workbook.close()

        except Exception as e:
            logger.error(f"Error finding subtable sheets: {e}")

        return subtable_sheets

    def _extract_subtables_from_sheet(self, file_path: str, sheet_name: str, reference_numbers: List[str]) -> List[SubtableItem]:
        """
        Extract subtables from a specific sheet by finding reference numbers and their associated data.
        """
        try:
            # Import the standalone extractor with correct path
            import sys
            import os
            # Add the backend directory to the path
            backend_dir = os.path.join(os.path.dirname(__file__), '..', '..')
            sys.path.insert(0, backend_dir)

            from excel_table_extractor_corrected import ExcelTableExtractorCorrected

            # Create extractor instance
            extractor = ExcelTableExtractorCorrected(file_path, sheet_name)

            # Find all reference numbers in this sheet
            reference_occurrences = []
            for row in range(1, extractor.worksheet.max_row + 1):
                for col in range(1, extractor.worksheet.max_column + 1):
                    cell_value = extractor.get_cell_value(row, col)
                    if cell_value and isinstance(cell_value, str):
                        for ref in reference_numbers:
                            if ref in cell_value:
                                reference_occurrences.append((row, col, ref))
                                logger.info(
                                    f"Found reference '{ref}' at row {row}, col {col} in sheet {sheet_name}")

            logger.info(
                f"Found {len(reference_occurrences)} reference occurrences in sheet {sheet_name}")

            # Extract subtables for each reference
            all_subtable_items = []
            for row, col, ref_number in reference_occurrences:
                # Find the header row (usually the next row after reference)
                header_row = self._find_header_row(extractor, row)
                if header_row:
                    logger.info(
                        f"Processing subtable for reference '{ref_number}' at row {row}")
                    logger.info(
                        f"Found header row {header_row} for reference '{ref_number}': '{extractor.get_cell_value(header_row, 2)}'")

                    # Extract data rows for this reference
                    logger.info(
                        f"Processing subtable for reference '{ref_number}' at row {row}")

                    # Find header row for this reference
                    header_row = self._find_header_row(extractor, row)
                    if header_row:
                        logger.info(
                            f"Found header row {header_row} for reference '{ref_number}': ''")

                        # Extract data rows using dynamic column detection
                        data_rows = self._extract_subtable_data_rows(
                            extractor, header_row, ref_number)

                        # Convert each data row to SubtableItem
                        for row_data in data_rows:
                            try:
                                logger.debug(
                                    f"About to create SubtableItem with dict: {row_data}")

                                # Create SubtableItem directly from the structured data
                                item = SubtableItem(
                                    item_key=row_data['item_name'],
                                    raw_fields=row_data['raw_fields'],
                                    quantity=row_data['quantity'],
                                    unit=row_data['unit'],
                                    source="Excel",
                                    reference_number=row_data['reference_number'],
                                    sheet_name=sheet_name
                                )

                                all_subtable_items.append(item)

                            except Exception as e:
                                logger.error(
                                    f"Error creating SubtableItem from row {row_data}: {e}")
                                continue
                    else:
                        logger.warning(
                            f"No header found for reference '{ref_number}' at row {row}")

            logger.info(
                f"Extracted {len(all_subtable_items)} items from {sheet_name}")
            return all_subtable_items

        except Exception as e:
            logger.error(
                f"Error extracting subtables from sheet {sheet_name}: {e}")
            return []

    def _is_table_title(self, item_name: str) -> bool:
        """
        Check if an item name is a table title that should be skipped.
        Only filter obvious structural elements, not valid items.
        """
        if not item_name or not item_name.strip():
            return False

        item_name = item_name.strip()

        # Filter obvious structural elements
        structural_elements = [
            "＊＊＊合計＊＊＊",
            "計",
            "諸雑費",
            "諸雑費(率+まるめ)",
            "諸雑費(まるめ)",
            "合計",
            "小計"
        ]

        if item_name in structural_elements:
            return True

        # Filter pure table titles without item descriptions
        # Examples: "30組当り明細書", "10m当り明細書" (without item name)
        import re
        if re.match(r'^\d+[a-zA-Zぁ-んァ-ヶー一-龯]*\s*当り明細書$', item_name):
            return True

        # Filter pure "明細書" patterns without item descriptions
        if item_name == "明細書" or item_name.endswith("明細書") and len(item_name) < 10:
            return True

        # Keep items with actual descriptions + "当り" (like "下地処理(表面含侵) 1m2当り明細書")
        # Keep items with actual descriptions + "当り" (like "運転手(特殊) 8時間当り")
        return False

    def _is_table_end(self, extractor, row: int) -> bool:
        """
        Check if we've reached the end of a table by looking for solid lines or empty rows.
        """
        # Don't stop on single empty rows - check multiple consecutive empty rows
        empty_row_count = 0
        for check_row in range(row, min(row + 3, extractor.worksheet.max_row + 1)):
            if self._is_row_empty(extractor, check_row):
                empty_row_count += 1
            else:
                break

        # Only consider it table end if we have 3+ consecutive empty rows
        if empty_row_count >= 3:
            return True

        # Check for solid line boundaries (thick borders) - but be more lenient
        try:
            for col in range(1, min(extractor.worksheet.max_column + 1, 10)):
                cell = extractor.worksheet.cell(row=row, column=col)
                if cell.border and cell.border.bottom:
                    # Check if it's a thick border (table boundary)
                    if hasattr(cell.border.bottom, 'style') and cell.border.bottom.style in ['thick']:
                        return True
        except:
            pass

        return False

    def _find_column_positions(self, extractor, header_row: int) -> Dict[str, int]:
        """
        Find column positions for name, unit, and quantity by looking for header patterns.
        Returns 0-based column indices.
        """
        positions = {}

        # Search for header patterns in the header row
        for col in range(1, min(extractor.worksheet.max_column + 1, 15)):
            cell_value = extractor.get_cell_value(header_row, col)
            if cell_value and isinstance(cell_value, str):
                cell_value = cell_value.strip()

                # Look for name/description column
                if any(pattern in cell_value for pattern in ['名称', '規格', '注釈']):
                    positions['name'] = col - 1  # Convert to 0-based
                    logger.debug(
                        f"Found name column at {col} (0-based: {col-1})")

                # Look for unit column
                elif any(pattern in cell_value for pattern in ['単位', '単 位']):
                    positions['unit'] = col - 1  # Convert to 0-based
                    logger.debug(
                        f"Found unit column at {col} (0-based: {col-1})")

                # Look for quantity column
                elif any(pattern in cell_value for pattern in ['数量', '数 量']):
                    positions['quantity'] = col - 1  # Convert to 0-based
                    logger.debug(
                        f"Found quantity column at {col} (0-based: {col-1})")

        # Override positions based on sheet type
        sheet_name = extractor.worksheet.title
        if "51明細書" in sheet_name:
            # Sheet 2 structure: name in col 2, unit in col 5, quantity in col 6
            positions['name'] = 1      # Column 2 (0-based: 1)
            positions['unit'] = 4      # Column 5 (0-based: 4)
            positions['quantity'] = 5  # Column 6 (0-based: 5)
            logger.debug(
                f"Sheet 2 override - Name: col 2, Unit: col 5, Quantity: col 6")
        elif "55単価表" in sheet_name:
            # Sheet 3 structure: name in col 2, unit in col 12, quantity in col 10
            positions['name'] = 1      # Column 2 (0-based: 1)
            positions['unit'] = 11     # Column 12 (0-based: 11)
            positions['quantity'] = 9  # Column 10 (0-based: 9)
            logger.debug(
                f"Sheet 3 override - Name: col 2, Unit: col 12, Quantity: col 10")

        logger.debug(f"Final column positions: {positions}")
        return positions

    def _extract_subtable_data_rows(self, extractor, header_row: int, reference_number: str) -> List[Dict[str, Any]]:
        """
        Extract data rows from a subtable using dynamic column detection and logical row merging.
        """
        # Find column positions dynamically
        column_positions = self._find_column_positions(extractor, header_row)

        # Default fallback positions if dynamic detection fails
        # Default to column 2 (0-based: 1)
        name_col = column_positions.get('name', 1)
        # Default to column 5 (0-based: 4)
        unit_col = column_positions.get('unit', 4)
        # Default to column 6 (0-based: 5)
        quantity_col = column_positions.get('quantity', 5)

        # Debug log the column assignments
        logger.info(
            f"Using columns - Name: {name_col+1}, Unit: {unit_col+1}, Quantity: {quantity_col+1}")

        # Extract logical rows (combining multi-row items between dotted lines)
        logical_rows = self._extract_logical_rows(
            extractor, header_row, name_col, unit_col, quantity_col, reference_number)

        data_rows = []
        for logical_row in logical_rows:
            # Create row data from logical row
            row_data = {
                'item_name': logical_row['combined_name'],
                'unit': logical_row['combined_unit'],
                'quantity': logical_row['combined_quantity'],
                'reference_number': reference_number,
                'raw_fields': {
                    '名称・規格': logical_row['combined_name'],
                    '単位': logical_row['combined_unit'],
                    '数量': str(logical_row['combined_quantity']) if logical_row['combined_quantity'] else '',
                    '摘要': reference_number
                }
            }
            data_rows.append(row_data)

        return data_rows

    def _extract_logical_rows(self, extractor, header_row: int, name_col: int, unit_col: int, quantity_col: int, reference_number: str) -> List[Dict[str, Any]]:
        """
        Extract logical rows by detecting item name + unit/quantity pairs.
        Each logical row consists of:
        1. A row with item name (in name_col)
        2. The next row with unit/quantity data (in unit_col/quantity_col)
        """
        logical_rows = []
        current_row = header_row + 1
        max_rows_to_check = 200  # Increased to capture more items
        rows_checked = 0

        while current_row <= extractor.worksheet.max_row and rows_checked < max_rows_to_check:
            rows_checked += 1

            # Check if we've reached the end of the table
            if self._is_table_end(extractor, current_row):
                break

            # Look for item names in the name column
            item_name = extractor.get_cell_value(
                current_row, name_col + 1)  # Convert to 1-based

            if item_name and isinstance(item_name, str):
                item_name_str = item_name.strip()

                # Skip meaningless content and table titles
                if not self._is_obvious_table_title(item_name_str) and not self._is_meaningless_content(item_name_str):
                    # Look for unit/quantity data in the next few rows
                    item_data = self._extract_item_with_data(
                        extractor, current_row, item_name_str, name_col, unit_col, quantity_col)

                    # Create logical row data with expected field names
                    logical_row = {
                        'combined_name': item_data['name'],
                        'combined_unit': item_data['unit'],
                        'combined_quantity': item_data['quantity'],
                        'reference_number': reference_number
                    }
                    logical_rows.append(logical_row)

                    # Skip to the end of this logical row
                    current_row = item_data['end_row'] + 1
                else:
                    current_row += 1
            else:
                current_row += 1

        return logical_rows

    def _extract_item_with_data(self, extractor, item_row: int, item_name: str, name_col: int, unit_col: int, quantity_col: int) -> Dict[str, Any]:
        """
        Extract a logical item starting from the item name row, looking for unit/quantity data in subsequent rows.
        """
        combined_name = item_name
        combined_unit = ""
        combined_quantity = 0.0
        end_row = item_row

        # Look for continued name parts horizontally across columns in the same row
        # Start from the next column after the main name column until the unit column
        for col in range(name_col + 1, unit_col):
            cell_value = extractor.get_cell_value(
                item_row, col + 1)  # Convert to 1-based
            if cell_value and isinstance(cell_value, str) and cell_value.strip():
                combined_name += " " + cell_value.strip()

        # Look for unit/quantity data in the next few rows (typically 1-3 rows after item name)
        for check_row in range(item_row + 1, min(item_row + 5, extractor.worksheet.max_row + 1)):
            # Check for unit and quantity in the current row
            unit_value = extractor.get_cell_value(
                check_row, unit_col + 1)  # Convert to 1-based
            quantity_value = extractor.get_cell_value(
                check_row, quantity_col + 1)  # Convert to 1-based

            # Also check for continued name parts in subsequent rows (same columns as above)
            for col in range(name_col + 1, unit_col):
                cell_value = extractor.get_cell_value(
                    check_row, col + 1)  # Convert to 1-based
                if cell_value and isinstance(cell_value, str) and cell_value.strip():
                    combined_name += " " + cell_value.strip()

            # If we found unit or quantity data, use it and stop
            if (unit_value and isinstance(unit_value, str) and unit_value.strip()) or \
               (quantity_value and str(quantity_value).strip()):

                if unit_value and isinstance(unit_value, str):
                    combined_unit = unit_value.strip()

                if quantity_value:
                    try:
                        combined_quantity = float(str(quantity_value))
                    except (ValueError, TypeError):
                        combined_quantity = 0.0

                end_row = check_row
                break

        return {
            'name': combined_name.strip(),
            'unit': combined_unit,
            'quantity': combined_quantity,
            'end_row': end_row
        }

    def _looks_like_new_item(self, extractor, row: int, name_col: int, unit_col: int, quantity_col: int) -> bool:
        """
        Check if a row looks like the start of a new logical item.
        """
        # Check if there's an item name in the name column
        item_name = extractor.get_cell_value(row, name_col + 1)
        if item_name and str(item_name).strip():
            item_name_str = str(item_name).strip()
            if not self._is_meaningless_content(item_name_str) and not self._is_obvious_table_title(item_name_str):
                # Check if this looks like a new item (has unit/quantity nearby)
                has_nearby_data = False
                for nearby_row in range(row, min(row + 3, extractor.worksheet.max_row + 1)):
                    nearby_unit = extractor.get_cell_value(
                        nearby_row, unit_col + 1)
                    nearby_quantity = extractor.get_cell_value(
                        nearby_row, quantity_col + 1)
                    if (nearby_unit and str(nearby_unit).strip()) or (nearby_quantity and str(nearby_quantity).strip()):
                        has_nearby_data = True
                        break

                return has_nearby_data

        return False

    def _is_meaningless_content(self, content: str) -> bool:
        """
        Check if content is meaningless (like spaces, dashes, etc.)
        """
        if not content:
            return True

        content = content.strip()
        if not content:
            return True

        # Skip content that's mostly whitespace or formatting characters
        meaningless_patterns = [
            '　',  # Full-width space
            ' ',   # Regular space
            '\t',  # Tab
            '\n',  # Newline
            '－',  # Dash
            '-',   # Hyphen
            '＝',  # Full-width equals
            '=',   # Equals
        ]

        # If content is only meaningless characters, skip it
        cleaned = content
        for pattern in meaningless_patterns:
            cleaned = cleaned.replace(pattern, '')

        return len(cleaned) == 0

    def _has_dotted_line(self, extractor, row: int) -> bool:
        """
        Check if a row contains dotted lines (horizontal separators).
        Enhanced to be more accurate for logical row boundaries.
        """
        # Check multiple columns for dotted line patterns
        dotted_line_found = False

        for col in range(1, min(extractor.worksheet.max_column + 1, 10)):
            cell_value = extractor.get_cell_value(row, col)
            if cell_value and isinstance(cell_value, str):
                cell_str = cell_value.strip()

                # Common dotted line patterns in Excel
                dotted_patterns = [
                    '─', '━', '＝', '－', '----', '====',
                    '・・・', '...', '‥‥', '……', '___'
                ]

                # Check if the cell contains primarily dotted line characters
                if len(cell_str) > 1:  # Must be at least 2 characters
                    for pattern in dotted_patterns:
                        if pattern in cell_str:
                            # Check if this is primarily a dotted line (not just containing the pattern)
                            pattern_count = cell_str.count(pattern)
                            if pattern_count > 0 and len(cell_str.replace(pattern, '').strip()) < len(cell_str) * 0.5:
                                dotted_line_found = True
                                break

                    if dotted_line_found:
                        break

        return dotted_line_found

    def _get_name_col_1based(self, extractor, row: int) -> int:
        """
        Get the 1-based column number for the name column (usually column 2).
        """
        return 2  # Most commonly column 2 in this Excel format

    def _is_definitive_table_end(self, extractor, row: int) -> bool:
        """
        Check for definitive table end - only stop on clear boundaries like next reference number.
        """
        # Check if we've reached the next reference number
        for col in range(1, min(extractor.worksheet.max_column + 1, 5)):
            cell_value = extractor.get_cell_value(row, col)
            if cell_value and isinstance(cell_value, str):
                cell_str = cell_value.strip()
                # If we find another reference number, this is definitely the end
                if '単' in cell_str and '号' in cell_str:
                    import re
                    if re.search(r'単\d+号', cell_str):
                        return True

        # Check for many consecutive empty rows (more than 10)
        empty_row_count = 0
        for check_row in range(row, min(row + 15, extractor.worksheet.max_row + 1)):
            if self._is_row_empty(extractor, check_row):
                empty_row_count += 1
            else:
                break

        # Only consider it table end if we have 10+ consecutive empty rows
        if empty_row_count >= 10:
            return True

        return False

    def _is_obvious_table_title(self, item_name: str) -> bool:
        """
        Check if this is an obvious table title that should be skipped.
        More permissive than the original _is_table_title method.
        """
        if not item_name:
            return False

        # Only skip very obvious table titles
        skip_patterns = [
            '明細書',
            '当り明細書',
            '計算書',
            '積算書'
        ]

        # Only skip if the item name ends with these patterns (not contains)
        for pattern in skip_patterns:
            if item_name.endswith(pattern):
                return True

        return False

    def _is_row_empty(self, extractor, row: int) -> bool:
        """Check if a row is empty (no content in any cell)"""
        for col in range(1, extractor.worksheet.max_column + 1):
            cell_value = extractor.get_cell_value(row, col)
            if cell_value and str(cell_value).strip():
                return False
        return True

    def _find_subtable_end(self, extractor, header_row: int) -> int:
        """
        Find the end row of the subtable: the next reference/header or end of sheet.
        """
        ws = extractor.worksheet
        max_row = ws.max_row
        # Look for the next reference/header row (cell in column 2 matching '単\d+号')
        for row in range(header_row + 1, max_row + 1):
            cell_value = ws.cell(row=row, column=2).value
            if cell_value and isinstance(cell_value, str):
                # If this row looks like a reference (e.g., '単12号'), treat as end
                import re
                if re.search(r'単\d+号', cell_value):
                    return row - 1  # End at the row before the next reference
        return max_row  # If no more references, end at last row

    def _extract_using_standalone_logic(self, file_path: str, sheet_name: str) -> List[TenderItem]:
        """
        Use the standalone excel_table_extractor_corrected.py logic to extract data.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to extract from

        Returns:
            List of TenderItem objects
        """
        try:
            logger.info(
                f"Starting standalone extraction from {file_path}, sheet: {sheet_name}")

            # Import the standalone extractor with correct path
            import sys
            import os
            # Add the backend directory to the path
            backend_dir = os.path.join(os.path.dirname(__file__), '..', '..')
            sys.path.insert(0, backend_dir)

            from excel_table_extractor_corrected import ExcelTableExtractorCorrected

            # Create extractor instance
            extractor = ExcelTableExtractorCorrected(file_path, sheet_name)

            # Extract all tables
            tables = extractor.extract_all_tables()

            logger.info(f"Standalone extractor found {len(tables)} tables")

            # Convert tables to TenderItem format
            tender_items = []

            for table_idx, table in enumerate(tables):
                logger.info(
                    f"Processing table {table_idx + 1} with {len(table.get('data_rows', []))} data rows")

                # Convert data rows to TenderItems
                table_items = self._convert_table_data_to_tender_items(
                    table, sheet_name, table_idx)
                tender_items.extend(table_items)

            logger.info(
                f"Total items extracted by standalone logic: {len(tender_items)}")

            # Debug: Print first few items to see their units
            for i, item in enumerate(tender_items[:5]):
                logger.info(
                    f"Item {i+1}: '{item.item_key}' | Unit: '{item.unit}' | Quantity: {item.quantity}")

            return tender_items

        except ImportError as e:
            logger.error(f"Could not import standalone extractor: {e}")
            raise
        except Exception as e:
            logger.error(f"Error in standalone extraction: {e}")
            raise

    def _convert_table_data_to_tender_items(self, table: Dict[str, Any], sheet_name: str, table_idx: int) -> List[TenderItem]:
        """
        Convert extracted table data to TenderItem format.

        Args:
            table: Table data from standalone extractor
            sheet_name: Name of the source sheet
            table_idx: Index of the table

        Returns:
            List of TenderItem objects
        """
        tender_items = []
        data_rows = table.get('data_rows', [])

        logger.info(
            f"Converting {len(data_rows)} data rows from table {table_idx + 1}")

        for row_idx, row_data in enumerate(data_rows):
            try:
                # Extract data from row (assuming standard column order)
                # [費目/工種/種別/細別/規格, 単位, 数量, 単価, 金額, 摘要]
                item_name = row_data[0] if len(row_data) > 0 else ""
                unit = row_data[1] if len(row_data) > 1 else ""
                quantity_str = row_data[2] if len(row_data) > 2 else ""
                unit_price_str = row_data[3] if len(row_data) > 3 else ""
                amount_str = row_data[4] if len(row_data) > 4 else ""
                remarks = row_data[5] if len(row_data) > 5 else ""

                # Debug: Log the raw row data for the target item
                if "補強部材取付工" in item_name:
                    logger.info(f"FOUND TARGET ITEM in row {row_idx}:")
                    logger.info(f"  Raw row data: {row_data}")
                    logger.info(f"  Item name: '{item_name}'")
                    logger.info(f"  Unit: '{unit}'")
                    logger.info(f"  Quantity: '{quantity_str}'")

                # Skip empty rows
                if not item_name.strip():
                    continue

                # Convert quantity to float
                quantity = 0.0
                try:
                    if quantity_str.strip():
                        # Remove commas and convert to float
                        clean_qty = quantity_str.replace(
                            ',', '').replace(' ', '').replace('　', '')
                        quantity = float(clean_qty)
                except (ValueError, TypeError):
                    quantity = 0.0

                # Create raw fields dictionary
                raw_fields = {
                    "工事区分・工種・種別・細別": item_name,
                    "規格": item_name,  # Use item name as specification
                }

                if unit:
                    raw_fields["単位"] = unit
                if quantity_str:
                    raw_fields["数量"] = quantity_str
                if unit_price_str:
                    raw_fields["単価"] = unit_price_str
                if amount_str:
                    raw_fields["金額"] = amount_str
                if remarks:
                    raw_fields["摘要"] = remarks

                # Create TenderItem
                tender_item = TenderItem(
                    item_key=item_name,
                    raw_fields=raw_fields,
                    quantity=quantity,
                    unit=unit,
                    source="Excel",
                    page_number=None
                )

                tender_items.append(tender_item)

                logger.debug(
                    f"Created TenderItem: {item_name} | {quantity} {unit}")

            except Exception as e:
                logger.warning(
                    f"Error converting row {row_idx} to TenderItem: {e}")
                continue

        logger.info(
            f"Successfully converted {len(tender_items)} items from table {table_idx + 1}")
        return tender_items

    def _convert_subtable_data_to_items(self, data_rows, ref_number, sheet_name):
        """
        Convert extracted subtable data rows to SubtableItem objects.
        """
        items = []
        for row in data_rows:
            # row: [item_name, unit, quantity, '', '', specification]
            item_key = (row[0] or "").strip()
            unit = (row[1] or "").strip() or None
            quantity_str = (row[2] or "").strip()
            try:
                quantity = float(quantity_str.replace(
                    ',', '')) if quantity_str else 0.0
            except Exception:
                quantity = 0.0
            specification = (row[5] or "").strip() if len(row) > 5 else ""
            raw_fields = {
                "名称・規格": item_key,
                "単位": unit or "",
                "単数": quantity_str,
                "摘要": specification,
            }
            item_dict = {
                "item_key": item_key,
                "raw_fields": raw_fields,
                "quantity": quantity,
                "unit": unit,
                "source": "Excel",
                "reference_number": ref_number,
                "sheet_name": sheet_name,
            }
            print(
                f"DEBUG: About to create SubtableItem with dict: {item_dict}")
            try:
                item = SubtableItem(**item_dict)
                items.append(item)
            except Exception as e:
                print(f"ERROR: Failed to create SubtableItem: {e}")
        return items

    def _extract_using_original_logic(self, excel_buffer: BytesIO, sheet_name: str) -> List[TenderItem]:
        """
        Fallback method that uses the original Excel parser logic when standalone logic fails.
        This ensures we always get some results even if the standalone logic has issues.
        """
        try:
            from .excel_parser import ExcelParser
            excel_parser = ExcelParser()
            return excel_parser.extract_items_from_buffer_with_sheet(excel_buffer, sheet_name)
        except Exception as e:
            logger.error(f"Error in fallback extraction: {e}")
            return []

    def _parse_quantity(self, quantity_str: str) -> float:
        """
        Parses a quantity string (e.g., "1,234", "1 234", "1,234.56") into a float.
        Handles various formats including commas and spaces.
        """
        if not quantity_str:
            return 0.0

        # Remove commas and spaces
        clean_str = quantity_str.replace(
            ',', '').replace(' ', '').replace('　', '')

        # Try to convert to float
        try:
            return float(clean_str)
        except ValueError:
            # If it's not a valid number, return 0.0
            return 0.0


# Keep the old class for backward compatibility but mark as deprecated
class ExcelTableExtractorCorrected:
    """
    DEPRECATED: This class is kept for backward compatibility.
    Use the standalone excel_table_extractor_corrected.py file instead.
    """

    def __init__(self, file_path: str, sheet_name: str):
        logger.warning(
            "ExcelTableExtractorCorrected is deprecated. Use standalone extractor instead.")
        raise NotImplementedError(
            "Use the standalone excel_table_extractor_corrected.py file instead")
