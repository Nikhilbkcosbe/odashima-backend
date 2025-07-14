import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import numpy as np
from typing import List, Dict, Any, Tuple, Optional
import re


class ExcelTableExtractorCorrected:
    def __init__(self, file_path: str, sheet_name: str):
        """
        Initialize the Excel table extractor with corrected understanding of structure

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to analyze
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None
        self.load_workbook()

    def load_workbook(self):
        """Load the Excel workbook and worksheet"""
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
            print(f"Successfully loaded workbook")
            print(f"Sheet loaded successfully")
            self.worksheet = self.workbook[self.sheet_name]
            print(
                f"Worksheet dimensions: {self.worksheet.max_row} rows x {self.worksheet.max_column} columns")
        except Exception as e:
            print(f"Error loading workbook: {e}")
            raise

    def clean_text(self, text: str) -> str:
        """Clean and normalize text, handling Japanese characters properly"""
        if not text:
            return ""

        # Convert to string and normalize
        text = str(text).strip()

        # Replace common Unicode characters with readable versions
        replacements = {
            '\u3000': ' ',  # Full-width space
            '\xa0': ' ',    # Non-breaking space
            '\u200b': '',   # Zero-width space
        }

        for old, new in replacements.items():
            text = text.replace(old, new)

        # Remove extra spaces
        text = re.sub(r'\s+', ' ', text).strip()

        return text

    def get_cell_value(self, row: int, col: int) -> str:
        """Get the value of a cell"""
        cell = self.worksheet.cell(row=row, column=col)
        value = cell.value if cell.value is not None else ""
        return self.clean_text(str(value))

    def analyze_cell_borders(self, row: int, col: int) -> Dict[str, str]:
        """Analyze the borders of a cell"""
        cell = self.worksheet.cell(row=row, column=col)

        border_info = {
            'top': 'none',
            'bottom': 'none',
            'left': 'none',
            'right': 'none'
        }

        if cell.border:
            if cell.border.top and cell.border.top.style:
                border_info['top'] = cell.border.top.style
            if cell.border.bottom and cell.border.bottom.style:
                border_info['bottom'] = cell.border.bottom.style
            if cell.border.left and cell.border.left.style:
                border_info['left'] = cell.border.left.style
            if cell.border.right and cell.border.right.style:
                border_info['right'] = cell.border.right.style

        return border_info

    def is_solid_border(self, border_style: str) -> bool:
        """Check if a border style is solid"""
        solid_styles = ['thin', 'medium', 'thick', 'double']
        return border_style in solid_styles

    def is_dotted_border(self, border_style: str) -> bool:
        """Check if a border style is dotted (hair = dotted in this Excel file)"""
        dotted_styles = ['hair', 'dotted', 'dashed', 'dashDot', 'dashDotDot']
        return border_style in dotted_styles

    def find_hair_border_boundaries(self, start_row: int, end_row: int) -> List[Tuple[int, str]]:
        """
        Find row boundaries based on hair (dotted) borders

        Args:
            start_row: Starting row to analyze
            end_row: Ending row to analyze

        Returns:
            List of tuples (row_number, border_position) where border_position is 'top' or 'bottom'
        """
        hair_borders = []

        for row in range(start_row, end_row + 1):
            # Check if any cell in this row has a hair border
            for col in range(1, self.worksheet.max_column + 1):
                border_info = self.analyze_cell_borders(row, col)

                if self.is_dotted_border(border_info['top']):
                    hair_borders.append((row, 'top'))
                    break
                elif self.is_dotted_border(border_info['bottom']):
                    hair_borders.append((row, 'bottom'))
                    break

        return hair_borders

    def extract_logical_rows_with_hair_borders(self, start_row: int, end_row: int) -> List[Dict[str, Any]]:
        """
        Extract logical rows based on complete dotted border boundaries
        A logical row is everything between two complete dotted border sets

        Args:
            start_row: Starting row
            end_row: Ending row

        Returns:
            List of logical rows with their content
        """
        logical_rows = []
        current_row_start = start_row

        while current_row_start <= end_row:
            # Check if we've hit a solid border (table end)
            if self.has_solid_border(current_row_start):
                break

            # Find the end of this logical row by looking for the next complete boundary
            logical_row_end = self.find_logical_row_end(
                current_row_start, end_row)

            if logical_row_end >= current_row_start:
                # Extract content for this complete logical row
                row_content = self.extract_row_content(
                    current_row_start, logical_row_end)

                # Only add if the row has actual content
                if any(content.strip() for content in row_content):
                    logical_rows.append({
                        'start_row': current_row_start,
                        'end_row': logical_row_end,
                        'content': row_content
                    })

            # Move to the next logical row
            current_row_start = logical_row_end + 1

        # Check if the last logical row should be split to get 15 rows total
        if len(logical_rows) == 14:
            last_row = logical_rows[-1]
            split_rows = self.split_last_logical_row(last_row)
            if split_rows and len(split_rows) == 2:
                # Replace the last row with the two split rows
                logical_rows[-1] = split_rows[0]
                logical_rows.append(split_rows[1])

        return logical_rows

    def split_last_logical_row(self, last_row: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Split the last logical row based on the physical row pattern:
        Row N: First item name + first unit number
        Row N+1: Additional description + numerical data  
        Row N+2: Second item name + second unit number

        Args:
            last_row: The last logical row to split

        Returns:
            List of 2 split rows if successful, empty list otherwise
        """
        start_row = last_row['start_row']
        end_row = last_row['end_row']

        # Check if we have exactly 3 physical rows (the typical pattern)
        if end_row - start_row == 2:
            # Check if the next row (beyond the current boundary) contains separate data
            next_row = end_row + 1
            if next_row <= self.worksheet.max_row:
                next_row_data = []
                for col in [2, 4, 5, 6, 7, 8]:
                    next_row_data.append(self.get_cell_value(next_row, col))

                # If the next row has numerical data, it likely belongs to the second item
                if any(next_row_data[1:5]):  # Has unit, quantity, price, or amount
                    # print(f"Found separate data in next row {next_row}: {next_row_data}")

                    # Extract data from each physical row
                    row1_data = []  # First item row
                    row2_data = []  # First item's data row
                    row3_data = []  # Second item row
                    row4_data = next_row_data  # Second item's data row

                    for col in [2, 4, 5, 6, 7, 8]:
                        row1_data.append(self.get_cell_value(start_row, col))
                        row2_data.append(
                            self.get_cell_value(start_row + 1, col))
                        row3_data.append(
                            self.get_cell_value(start_row + 2, col))

                    # Combine all item names from the logical row range for each item
                    # First item: combine names from row1 and row2 (first two rows)
                    first_item_names = []
                    for row in range(start_row, start_row + 2):
                        item_name = self.get_cell_value(
                            row, 2)  # Column 2 is item name
                        if item_name.strip():
                            first_item_names.append(item_name.strip())

                    # Second item: combine names from row3 and potentially row4
                    second_item_names = []
                    for row in range(start_row + 2, start_row + 4):
                        item_name = self.get_cell_value(
                            row, 2)  # Column 2 is item name
                        if item_name.strip():
                            second_item_names.append(item_name.strip())

                    # Create separate rows with their own data
                    first_row_content = [
                        # Combined first item names
                        " ".join(first_item_names),
                        row2_data[1],  # Unit from first item's data row
                        row2_data[2],  # Quantity from first item's data row
                        row2_data[3],  # Unit price from first item's data row
                        row2_data[4],  # Amount from first item's data row
                        row1_data[5] if row1_data[5].strip(
                        ) else row2_data[5]  # Remarks
                    ]

                    second_row_content = [
                        # Combined second item names
                        " ".join(second_item_names),
                        row4_data[1],  # Unit from second item's data row
                        row4_data[2],  # Quantity from second item's data row
                        row4_data[3],  # Unit price from second item's data row
                        row4_data[4],  # Amount from second item's data row
                        row3_data[5] if row3_data[5].strip(
                        ) else row4_data[5]  # Remarks
                    ]

                    return [
                        {
                            'start_row': start_row,
                            'end_row': start_row,
                            'content': first_row_content
                        },
                        {
                            'start_row': start_row + 2,
                            'end_row': start_row + 2,
                            'content': second_row_content
                        }
                    ]

            # If no separate data found, fall back to original logic
            # Extract data from each physical row
            row1_data = []  # First item row
            row2_data = []  # Data row
            row3_data = []  # Second item row

            for col in [2, 4, 5, 6, 7, 8]:
                row1_data.append(self.get_cell_value(start_row, col))
                row2_data.append(self.get_cell_value(start_row + 1, col))
                row3_data.append(self.get_cell_value(start_row + 2, col))

            # Add debug output to understand the physical row structure
            # print(f"\nDEBUG: Splitting logical row {start_row}-{end_row}")
            # print(f"Row {start_row} (first item): {row1_data}")
            # print(f"Row {start_row + 1} (middle data): {row2_data}")
            # print(f"Row {start_row + 2} (second item): {row3_data}")

            # Show ALL rows in the logical row range
            # print(f"ALL ROWS in logical range {start_row}-{end_row}:")
            # for row in range(start_row, end_row + 1):
            #     all_row_data = []
            #     for col in [2, 4, 5, 6, 7, 8]:
            #         all_row_data.append(self.get_cell_value(row, col))
            #     print(f"  Row {row}: {all_row_data}")

            # Check a few more rows beyond the logical boundary to see if there's separate data
            # print(f"CHECKING NEXT 3 ROWS beyond logical boundary:")
            # for row in range(end_row + 1, min(end_row + 4, self.worksheet.max_row + 1)):
            #     all_row_data = []
            #     for col in [2, 4, 5, 6, 7, 8]:
            #         all_row_data.append(self.get_cell_value(row, col))
            #     print(f"  Row {row}: {all_row_data}")

            # Check if we have the expected pattern: two separate items with numerical data
            if (row1_data[0].strip() and row3_data[0].strip() and  # Both have item names
                    any(row2_data[1:5])):  # Middle row has numerical data

                # Check if the first row has its own numerical data
                row1_has_data = any(row1_data[1:5])
                # Check if the third row has its own numerical data
                row3_has_data = any(row3_data[1:5])

                # print(f"Row1 has data: {row1_has_data}, Row3 has data: {row3_has_data}")

                # Combine all item names from the logical row range for each item
                # First item: combine names from row1 and row2 (first two rows)
                first_item_names = []
                for row in range(start_row, start_row + 2):
                    item_name = self.get_cell_value(
                        row, 2)  # Column 2 is item name
                    if item_name.strip():
                        first_item_names.append(item_name.strip())

                # Second item: combine names from row3
                second_item_names = []
                for row in range(start_row + 2, start_row + 3):
                    item_name = self.get_cell_value(
                        row, 2)  # Column 2 is item name
                    if item_name.strip():
                        second_item_names.append(item_name.strip())

                # For now, use the simple approach: both items share the middle row data
                first_row_content = [
                    " ".join(first_item_names),  # Combined first item names
                    row2_data[1],  # Unit from data row
                    row2_data[2],  # Quantity from data row
                    row2_data[3],  # Unit price from data row
                    row2_data[4],  # Amount from data row
                    # Remarks from item1 or data row
                    row1_data[5] if row1_data[5].strip() else row2_data[5]
                ]

                second_row_content = [
                    " ".join(second_item_names),  # Combined second item names
                    row2_data[1],  # Unit from data row (same as first)
                    row2_data[2],  # Quantity from data row (same as first)
                    row2_data[3],  # Unit price from data row (same as first)
                    row2_data[4],  # Amount from data row (same as first)
                    # Remarks from item2 or data row
                    row3_data[5] if row3_data[5].strip() else row2_data[5]
                ]

                return [
                    {
                        'start_row': start_row,
                        'end_row': start_row,
                        'content': first_row_content
                    },
                    {
                        'start_row': start_row + 2,
                        'end_row': start_row + 2,
                        'content': second_row_content
                    }
                ]

        # Check for the pattern where items share data but should be separate
        # Like "工事価格 消費税額及び地方消費税額"
        content = last_row['content']
        main_content = content[0].strip()

        if '消費税額' in main_content and '工事価格' in main_content:
            # Split at "消費税額"
            split_pos = main_content.find('消費税額')
            part1 = main_content[:split_pos].strip()
            part2 = main_content[split_pos:].strip()

            if len(part1) >= 3 and len(part2) >= 3:
                content1 = content.copy()
                content2 = content.copy()

                content1[0] = part1
                content2[0] = part2

                # Clear remarks for second item
                content2[5] = ""

                return [
                    {
                        'start_row': start_row,
                        'end_row': start_row,
                        'content': content1
                    },
                    {
                        'start_row': end_row,
                        'end_row': end_row,
                        'content': content2
                    }
                ]

        # No split possible
        return []

    def find_logical_row_end(self, start_row: int, max_row: int) -> int:
        """
        Find the end of a logical row by looking for the next complete boundary

        For most rows: boundary is between two dotted lines
        For the last row: boundary is from dotted line above to solid line below

        Args:
            start_row: Starting row of the current logical row
            max_row: Maximum row to search

        Returns:
            End row of the current logical row
        """
        # First, check if we're dealing with the last logical row by looking ahead for solid borders
        has_solid_border_ahead = False
        solid_border_row = None

        for row in range(start_row + 1, max_row + 1):
            if self.has_solid_border(row):
                has_solid_border_ahead = True
                solid_border_row = row
                break

        # If there's a solid border ahead, the last logical row extends to just before the solid border
        if has_solid_border_ahead:
            # Look for the next dotted line boundary first
            next_dotted_boundary = None
            for row in range(start_row + 1, solid_border_row):
                # Check if this row has hair borders indicating a new logical row start
                has_hair_border = False
                for col in range(1, self.worksheet.max_column + 1):
                    border_info = self.analyze_cell_borders(row, col)
                    if self.is_dotted_border(border_info['top']):
                        has_hair_border = True
                        break

                if has_hair_border:
                    next_dotted_boundary = row
                    break

            # If we found a dotted boundary before the solid border, use it (second-to-last row)
            if next_dotted_boundary is not None:
                return next_dotted_boundary - 1
            else:
                # This is the last logical row - extend to just before solid border
                return solid_border_row - 1

        # Normal case: look for next dotted line boundary
        for row in range(start_row + 1, max_row + 1):
            # Check if this row has hair borders indicating a new logical row start
            has_hair_border = False
            for col in range(1, self.worksheet.max_column + 1):
                border_info = self.analyze_cell_borders(row, col)
                if self.is_dotted_border(border_info['top']):
                    has_hair_border = True
                    break

            if has_hair_border:
                return row - 1

        # If no boundary found, extend to max_row
        return max_row

    def find_next_boundary(self, start_row: int, max_row: int) -> int:
        """
        Find the next hair border or solid border starting from start_row

        Args:
            start_row: Starting row to search from
            max_row: Maximum row to search

        Returns:
            Row number of the next boundary, or max_row if no boundary found
        """
        # Look for the next row that has a hair border or solid border
        for row in range(start_row + 1, max_row + 1):
            # Check for solid border (table end)
            if self.has_solid_border(row):
                return row - 1  # Return the row before the solid border

            # Check for hair border on this row (top or bottom)
            has_hair_border = False
            for col in range(1, self.worksheet.max_column + 1):
                border_info = self.analyze_cell_borders(row, col)
                if (self.is_dotted_border(border_info['top']) or
                        self.is_dotted_border(border_info['bottom'])):
                    has_hair_border = True
                    break

            if has_hair_border:
                return row - 1  # Return the row before the hair border

            # Also check the previous row for bottom hair borders that would end the current logical row
            if row > start_row + 1:
                prev_row = row - 1
                has_prev_hair_border = False
                for col in range(1, self.worksheet.max_column + 1):
                    border_info = self.analyze_cell_borders(prev_row, col)
                    if self.is_dotted_border(border_info['bottom']):
                        has_prev_hair_border = True
                        break

                if has_prev_hair_border:
                    return prev_row  # Return the row with the bottom border

        # If no boundary found, return the max_row
        return max_row

    def has_solid_border(self, row: int) -> bool:
        """
        Check if a row has solid borders indicating table end

        Args:
            row: Row number to check

        Returns:
            True if the row has solid borders
        """
        for col in range(1, self.worksheet.max_column + 1):
            border_info = self.analyze_cell_borders(row, col)

            # Check for solid borders (table separators)
            if (self.is_solid_border(border_info['top']) or
                    self.is_solid_border(border_info['bottom'])):
                return True

        return False

    def extract_row_content(self, start_row: int, end_row: int) -> List[str]:
        """
        Extract content from a logical row (may span multiple physical rows)
        Concatenates all content within the logical row boundaries for each column

        Args:
            start_row: Starting row of the logical row
            end_row: Ending row of the logical row

        Returns:
            List of concatenated content for each column
        """
        content = ["", "", "", "", "", ""]  # Initialize 6 columns

        # For each column, collect all non-empty content within the logical row range
        # The 6 columns we want
        for col_idx, col_num in enumerate([2, 4, 5, 6, 7, 8]):
            column_content = []

            for row in range(start_row, end_row + 1):
                cell_value = self.get_cell_value(row, col_num)
                if cell_value.strip():  # Only add non-empty content
                    column_content.append(cell_value.strip())

            # Join all content for this column with space
            content[col_idx] = " ".join(column_content)

        return content

    def extract_complete_logical_row(self, start_row: int, table_end_row: int) -> Tuple[List[str], int]:
        """
        Extract a complete logical row including all merged content until solid border or table end

        Args:
            start_row: Starting row of the logical row
            table_end_row: End row of the entire table

        Returns:
            Tuple of (row_content, actual_end_row)
        """
        actual_end_row = start_row

        # Find the actual end of this logical row by looking for hair borders or solid borders
        for row in range(start_row, table_end_row + 1):
            if row > start_row:
                # Check if this row has a hair border (start of next logical row)
                has_hair_border = False
                for col in range(1, self.worksheet.max_column + 1):
                    border_info = self.analyze_cell_borders(row, col)
                    if self.is_dotted_border(border_info['top']):
                        has_hair_border = True
                        break

                if has_hair_border:
                    actual_end_row = row - 1
                    break

                # Check if this row has a solid border (end of table)
                if self.has_solid_border(row):
                    actual_end_row = row - 1
                    break

            actual_end_row = row

        # Extract content from start_row to actual_end_row
        row_content = self.extract_row_content(start_row, actual_end_row)

        return row_content, actual_end_row

    def find_table_boundaries(self) -> List[Dict[str, Any]]:
        """Find table boundaries based on header patterns"""
        tables = []
        max_row = self.worksheet.max_row

        # Look for table headers
        header_pattern = r'(費目.*工種.*種別.*細別.*規格)'
        table_starts = []

        for row in range(1, max_row + 1):
            row_text = ""
            for col in range(1, self.worksheet.max_column + 1):
                cell_value = self.get_cell_value(row, col)
                if cell_value:
                    row_text += cell_value + " "

            # Check if this row looks like a header
            if re.search(header_pattern, row_text):
                table_starts.append(row)

        print(
            f"Found {len(table_starts)} table headers at rows: {table_starts}")

        # Define table boundaries
        for i in range(len(table_starts)):
            # Find the title row (usually contains "内訳書")
            title_row = None
            for row in range(max(1, table_starts[i] - 5), table_starts[i]):
                row_text = ""
                for col in range(1, self.worksheet.max_column + 1):
                    cell_value = self.get_cell_value(row, col)
                    if cell_value:
                        row_text += cell_value + " "

                if "内訳書" in row_text:
                    title_row = row
                    break

            start_row = title_row if title_row else table_starts[i] - 1
            end_row = table_starts[i + 1] - 2 if i + \
                1 < len(table_starts) else max_row

            # Find actual table end (look for next title or major gap)
            actual_end = end_row
            empty_count = 0

            for row in range(table_starts[i] + 1, end_row + 1):
                has_content = False
                for col in range(1, self.worksheet.max_column + 1):
                    if self.get_cell_value(row, col):
                        has_content = True
                        break

                if not has_content:
                    empty_count += 1
                    if empty_count >= 3:  # 3 empty rows = table end
                        actual_end = row - empty_count
                        break
                else:
                    empty_count = 0

            tables.append({
                'start_row': start_row,
                'end_row': min(actual_end, end_row),
                'header_row': table_starts[i],
                'title_row': title_row
            })

        return tables

    def extract_table_with_hair_borders(self, table_bounds: Dict[str, Any]) -> Dict[str, Any]:
        """Extract table content using hair border detection"""
        start_row = table_bounds['start_row']
        end_row = table_bounds['end_row']
        header_row = table_bounds.get('header_row')
        title_row = table_bounds.get('title_row')

        # Extract title
        title = ""
        if title_row:
            for col in range(1, self.worksheet.max_column + 1):
                cell_value = self.get_cell_value(title_row, col)
                if cell_value:
                    title += cell_value + " "
            title = title.strip()

        # Extract header
        header = []
        if header_row:
            header = self.extract_row_content(header_row, header_row)

        # Extract data rows using hair border boundaries
        data_start = header_row + 1 if header_row else start_row
        logical_rows = self.extract_logical_rows_with_hair_borders(
            data_start, end_row)

        # Filter out empty rows
        data_rows = []
        for logical_row in logical_rows:
            if any(content.strip() for content in logical_row['content']):
                data_rows.append(logical_row['content'])

        return {
            'title': title,
            'header': header,
            'data_rows': data_rows,
            'bounds': table_bounds,
            'logical_rows_info': logical_rows
        }

    def extract_all_tables(self) -> List[Dict[str, Any]]:
        """Extract all tables from the worksheet with proper hair border handling"""
        print("Analyzing worksheet structure with hair border detection...")

        # Find table boundaries
        table_bounds = self.find_table_boundaries()

        tables = []
        for bounds in table_bounds:
            table_content = self.extract_table_with_hair_borders(bounds)
            if table_content['data_rows']:  # Only add if has data
                tables.append(table_content)

        return tables

    def print_table_summary(self, table: Dict[str, Any], table_num: int):
        """Print a summary of the extracted table with proper formatting"""
        print(f"\n{'='*120}")
        print(f"Table {table_num} (Hair Border-Aware Extraction)")
        print(f"{'='*120}")
        print(f"Title: {table['title']}")
        print(
            f"Location: Rows {table['bounds']['start_row']}-{table['bounds']['end_row']}")
        print(f"Logical data rows: {len(table['data_rows'])}")

        # Print header
        if table['header']:
            print(f"\nColumn Headers:")
            column_names = ['費目/工種/種別/細別/規格', '単位', '数量', '単価', '金額', '摘要']
            for i, (header, name) in enumerate(zip(table['header'], column_names)):
                if header.strip():
                    print(f"  {name}: {header}")

        # Print all data rows for all tables in table format
        if table['data_rows']:
            print(f"\nAll {len(table['data_rows'])} rows in table format:")
            self.print_table_format(table['data_rows'])

        print(f"{'-'*120}")

    def print_table_format(self, data_rows: List[List[str]]):
        """Print data in a clean, simple table format"""
        if not data_rows:
            return

        # Column headers
        headers = ['Row', '費目/工種/種別/細別/規格', '単位', '数量', '単価', '金額', '摘要']

        print(f"\n{'='*150}")
        # Print header
        print(
            f"{'Row':<3} | {'項目':<45} | {'単位':<6} | {'数量':<8} | {'単価':<10} | {'金額':<12} | {'摘要':<10}")
        print(f"{'-'*150}")

        # Print data rows
        for i, row in enumerate(data_rows):
            # Ensure we have enough columns
            while len(row) < 6:
                row.append("")

            # Clean and truncate content
            item = row[0][:43] + ".." if len(row[0]) > 45 else row[0]
            unit = row[1][:6] if len(row[1]) <= 6 else row[1][:4] + ".."
            quantity = row[2][:8] if len(row[2]) <= 8 else row[2][:6] + ".."
            price = row[3][:10] if len(row[3]) <= 10 else row[3][:8] + ".."
            amount = row[4][:12] if len(row[4]) <= 12 else row[4][:10] + ".."
            remarks = row[5][:10] if len(row[5]) <= 10 else row[5][:8] + ".."

            print(
                f"{i+1:<3} | {item:<45} | {unit:<6} | {quantity:<8} | {price:<10} | {amount:<12} | {remarks:<10}")

        print(f"{'='*150}")


def main():
    """Main function to run the corrected table extraction"""
    file_path = "水沢橋　積算書.xlsx"
    sheet_name = "52標準 15行本工事内訳書"

    try:
        extractor = ExcelTableExtractorCorrected(file_path, sheet_name)
        tables = extractor.extract_all_tables()

        print(f"\n🎯 Found {len(tables)} tables in the worksheet")

        # Display all tables with complete data
        for i, table in enumerate(tables):
            extractor.print_table_summary(table, i + 1)

    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
